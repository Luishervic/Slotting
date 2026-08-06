"""Intercambio simplificado del layout entre la aplicación y Excel.

El usuario administra cuatro hojas visibles: avance, tipos, un boceto editable
y el último resultado físico restringido. Las tablas normalizadas permanecen
ocultas como contrato técnico; la aplicación reconstruye las coordenadas,
asigna los SKU compatibles y regenera el resultado al reimportar el boceto.
"""
from __future__ import annotations

import io
import json
import math
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import (Alignment, Border, Font, PatternFill,
                             Protection, Side)
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from slotting.geometry import (normalizar_poligono, punto_en_poligono,
                               rectangulo_en_poligono)


AZUL = "12304A"
AZUL_2 = "0E7490"
VERDE = "059669"
MORADO = "7C3AED"
GRIS = "E2E8F0"
AMARILLO = "FEF3C7"
BLANCO = "FFFFFF"


LOCALIDAD_COLS = [
    "id_localidad", "codigo_wms", "tipo_codigo", "zona_layout",
    "x_m", "y_m", "ancho_m", "fondo_m", "alto_util_m", "orientacion",
    "abc_permitido", "departamento_permitido", "clase_permitida",
    "familia_permitida", "zona_fisica_permitida", "multisku", "activa",
    "notas", "estado_preparacion",
]


def _lista_texto(valor) -> str:
    if isinstance(valor, (list, tuple, set)):
        return ", ".join(str(v) for v in valor if str(v).strip())
    return "" if valor in (None, "") else str(valor)


def _booleano(valor, default: bool = False) -> bool:
    if isinstance(valor, bool):
        return valor
    if valor is None:
        return default
    return str(valor).strip().lower() in {"1", "true", "verdadero", "si", "sí", "x"}


def _encabezado(ws, fila: int, columnas: Iterable[str]) -> None:
    for col, valor in enumerate(columnas, 1):
        c = ws.cell(fila, col, valor)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.font = Font(color=BLANCO, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[fila].height = 30


def _tabla_estilo(ws, fila_header: int, fila_fin: int, ncols: int) -> None:
    borde = Border(bottom=Side(style="thin", color="CBD5E1"))
    for fila in ws.iter_rows(min_row=fila_header + 1, max_row=max(fila_header + 1, fila_fin),
                             min_col=1, max_col=ncols):
        for c in fila:
            c.border = borde
            c.alignment = Alignment(vertical="top", wrap_text=False)
    ws.auto_filter.ref = f"A{fila_header}:{get_column_letter(ncols)}{max(fila_header, fila_fin)}"
    ws.freeze_panes = f"A{fila_header + 1}"


def _tabla_excel(ws, nombre: str, ncols: int) -> None:
    """Convierte el rango usado en una tabla filtrable y auditable."""
    if ws.max_row < 2:
        return
    ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"
    # La tabla ya incorpora su propio AutoFilter. Mantener además el filtro
    # de hoja sobre el mismo rango produce dos filtros superpuestos; Excel
    # repara el archivo eliminando la tabla completa.
    ws.auto_filter.ref = None
    tabla = Table(displayName=nombre, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tabla)


def _crear_instrucciones(wb: Workbook, escala_m: float) -> None:
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    ws["A1"] = "Layout editable de localidades"
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL)
    ws["A1"].font = Font(color=BLANCO, bold=True, size=18)
    ws["A1"].alignment = Alignment(vertical="center")
    filas = [
        (4, "Orden recomendado", "1) Revisa Dashboard y Tipos_ubicacion. 2) Distribuye tipos en Mapa_preliminar. 3) Guarda y reimporta. 4) Descarga de nuevo para revisar Mapa_restringido."),
        (6, "Fuente de verdad", "Mapa_preliminar es la entrada editable. Mapa_restringido refleja el último resultado validado; las tablas de soporte permanecen ocultas."),
        (8, "Escala del mapa", f"Cada celda representa aproximadamente {escala_m:g} m × {escala_m:g} m. Sirve para comprobar proporciones, no para sustituir coordenadas."),
        (10, "Restricciones", "Vacío significa libre. Se pueden restringir localidades por ABC, departamento, clase, familia o zona física."),
        (12, "Códigos", "id_localidad es la llave técnica y no debe repetirse. En Relacion_SKU_Localidad puede repetirse para permitir que una localidad multi-SKU reciba varios códigos."),
        (14, "Coordenadas", "X/Y corresponden a la esquina inferior izquierda; ancho y fondo están en metros. La orientación es informativa: cambia ancho/fondo para rotar físicamente."),
        (16, "Antes de aplicar", "Reimportar ejecuta de nuevo la validación geométrica. Un error bloquea la aplicación; una advertencia permite continuar."),
    ]
    for fila, titulo, texto in filas:
        ws[f"A{fila}"] = titulo
        ws[f"A{fila}"].font = Font(bold=True, color=AZUL)
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=8)
        ws.cell(fila, 2, texto).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[fila].height = 34
    ws.column_dimensions["A"].width = 22
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16


def _crear_tipos(wb: Workbook, tipos: list[dict]) -> None:
    ws = wb.create_sheet("Tipos_ubicacion")
    headers = ["tipo_codigo", "nombre", "zona_fisica", "estructura",
               "ancho_m", "fondo_m", "alto_util_m", "rotacion_producto",
               "estado_medidas"]
    _encabezado(ws, 1, headers)
    for i, t in enumerate(tipos, 2):
        ws.append([
            t.get("codigo"), t.get("tipo"), t.get("zona_fisica"),
            t.get("tipo_estructura"), t.get("w"), t.get("d"), t.get("h"),
            t.get("orientacion_producto"), t.get("estado_medidas"),
        ])
        for col in range(5, 8):
            ws.cell(i, col).number_format = "0.00"
    _tabla_estilo(ws, 1, ws.max_row, len(headers))
    _tabla_excel(ws, "tblTiposUbicacion", len(headers))
    widths = [18, 30, 22, 14, 12, 12, 14, 24, 18]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.protection.sheet = True
    ws.protection.enable()


def _crear_zonas(wb: Workbook, zonas: list[dict]) -> None:
    ws = wb.create_sheet("Zonas")
    headers = ["zona", "prioridad", "forma", "vertices_json", "x_m", "y_m",
               "ancho_m", "fondo_m", "zona_fisica", "estructura",
               "pasillos", "ancho_pasillo_m", "orientacion", "margen_m",
               "tipos_permitidos", "departamentos", "clases", "familias", "abc"]
    _encabezado(ws, 1, headers)
    for z in zonas:
        poly = z.get("poligono") or []
        ws.append([
            z.get("nombre"), z.get("prioridad", 1),
            "Polígono" if poly else "Rectángulo",
            json.dumps(poly, ensure_ascii=False) if poly else "",
            z.get("x"), z.get("y"), z.get("w"), z.get("d"),
            z.get("zona_fisica", ""), z.get("tipo_estructura", "Automática"),
            z.get("modo_pasillo", "auto"), z.get("pasillo_m", 3.5),
            z.get("orientacion", "automatica"), z.get("margen_m", 0.5),
            _lista_texto(z.get("tipos")), _lista_texto(z.get("departamentos")),
            _lista_texto(z.get("clases")), _lista_texto(z.get("familias")),
            _lista_texto(z.get("abc")),
        ])
    _tabla_estilo(ws, 1, ws.max_row, len(headers))
    _tabla_excel(ws, "tblZonas", len(headers))
    for col in range(5, 15):
        ws.cell(1, col).alignment = Alignment(horizontal="center", wrap_text=True)
    for i, width in enumerate([22, 10, 12, 42, 10, 10, 12, 12, 20, 14, 12,
                               16, 15, 12, 28, 26, 24, 24, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _crear_localidades(wb: Workbook, slots: list[dict]) -> None:
    ws = wb.create_sheet("Localidades")
    _encabezado(ws, 1, LOCALIDAD_COLS)
    for i, s in enumerate(slots, 2):
        ws.append([
            s.get("id"), s.get("codigo_wms", ""), s.get("tipo_codigo", ""),
            s.get("zona_layout", ""), s.get("x"), s.get("y"), s.get("w"),
            s.get("d"), s.get("altura_util_nivel_m", s.get("h")),
            s.get("orientacion", ""), _lista_texto(s.get("clase_abc_reservada")),
            _lista_texto(s.get("departamento_reservado")),
            _lista_texto(s.get("clase_comercial_reservada")),
            _lista_texto(s.get("familia_reservada")),
            _lista_texto(s.get("zona_fisica_reservada")),
            bool(s.get("multisku", False)), bool(s.get("activa", True)),
            s.get("notas", ""),
            f'=IF(OR(A{i}="",C{i}="",D{i}="",E{i}="",F{i}="",G{i}="",H{i}=""),"PENDIENTE","PREPARADA")',
        ])
    _tabla_estilo(ws, 1, ws.max_row, len(LOCALIDAD_COLS))
    _tabla_excel(ws, "tblLocalidades", len(LOCALIDAD_COLS))
    for row in range(2, ws.max_row + 1):
        for col in range(5, 10):
            ws.cell(row, col).number_format = "0.00"
    widths = [18, 20, 18, 22, 10, 10, 12, 12, 14, 14, 14, 28, 26, 26, 26,
              11, 10, 32, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    # Entradas editables visibles; las coordenadas también se pueden editar.
    for row in ws.iter_rows(min_row=2, max_row=max(2, ws.max_row), min_col=2, max_col=18):
        for c in row:
            c.fill = PatternFill("solid", fgColor="FFFDF0")
    ws.conditional_formatting.add(
        f"S2:S{max(500, ws.max_row)}",
        FormulaRule(formula=['S2="PREPARADA"'],
                    fill=PatternFill("solid", fgColor="DCFCE7")))


def _crear_relacion(wb: Workbook, slots: list[dict],
                    relaciones: list[dict] | None = None) -> int:
    ws = wb.create_sheet("Relacion_SKU_Localidad")
    headers = ["id_localidad", "codigo_wms", "tipo_codigo", "zona_sugerida",
               "sku_asignado", "estado_geometria", "estado_designacion",
               "observaciones"]
    _encabezado(ws, 1, headers)
    loc_fin = max(500, len(slots) + 101)
    por_localidad: dict[str, list[str]] = {}
    for relacion in relaciones or []:
        uid = str(relacion.get("id_localidad") or "").strip()
        sku = str(relacion.get("sku") or "").strip()
        if uid and sku:
            por_localidad.setdefault(uid, []).append(sku)
    filas = [(s, sku) for s in slots
             for sku in (por_localidad.get(str(s.get("id"))) or [""])]
    for i, (s, sku) in enumerate(filas, 2):
        ws.append([
            s.get("id"),
            f'=IFERROR(INDEX(\'Localidades\'!$B$2:$B${loc_fin},MATCH(A{i},\'Localidades\'!$A$2:$A${loc_fin},0)),"")',
            s.get("tipo_codigo", ""), s.get("zona_layout", ""), sku,
            f'=IFERROR(INDEX(\'Localidades\'!$S$2:$S${loc_fin},MATCH(A{i},\'Localidades\'!$A$2:$A${loc_fin},0)),"SIN LOCALIDAD")',
            f'=IF(E{i}="","PENDIENTE SKU",IF(F{i}="PREPARADA","UBICADO","FALTA GEOMETRÍA"))',
            "",
        ])
    _tabla_estilo(ws, 1, ws.max_row, len(headers))
    _tabla_excel(ws, "tblRelacionSkuLocalidad", len(headers))
    for col, width in enumerate([18, 20, 18, 22, 20, 20, 22, 34], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2, max_row=max(2, ws.max_row),
                            min_col=5, max_col=5):
        row[0].fill = PatternFill("solid", fgColor="FFFDF0")
    ws.conditional_formatting.add(
        f"G2:G{max(500, ws.max_row)}",
        FormulaRule(formula=['G2="UBICADO"'],
                    fill=PatternFill("solid", fgColor="DCFCE7")))
    ws.conditional_formatting.add(
        f"G2:G{max(500, ws.max_row)}",
        FormulaRule(formula=['G2<>"UBICADO"'],
                    fill=PatternFill("solid", fgColor="FEF3C7")))
    return max(500, ws.max_row + 100)


def _crear_sku_designar(wb: Workbook, requerimientos: list[dict],
                        relacion_fin: int) -> int:
    ws = wb.create_sheet("SKU_por_designar")
    headers = ["sku", "descripcion", "unidades", "abc", "departamento",
               "clase", "familia", "zona_fisica", "tipo_codigo",
               "capacidad_localidad", "localidades_necesarias",
               "localidades_ubicadas", "pendientes", "estado"]
    _encabezado(ws, 1, headers)
    for i, r in enumerate(requerimientos, 2):
        ws.append([
            r.get("sku"), r.get("descripcion", ""), r.get("unidades", 0),
            r.get("abc", ""), r.get("departamento", ""), r.get("clase", ""),
            r.get("familia", ""), r.get("zona_fisica", ""),
            r.get("tipo_codigo", ""), r.get("capacidad_localidad", 0),
            r.get("localidades_necesarias", 0),
            f'=COUNTIFS(\'Relacion_SKU_Localidad\'!$E$2:$E${relacion_fin},A{i},\'Relacion_SKU_Localidad\'!$G$2:$G${relacion_fin},"UBICADO")',
            f'=MAX(0,K{i}-L{i})',
            f'=IF(L{i}=0,"PENDIENTE",IF(M{i}>0,"PARCIAL","UBICADO"))',
        ])
    _tabla_estilo(ws, 1, ws.max_row, len(headers))
    _tabla_excel(ws, "tblSkuPorDesignar", len(headers))
    for col, width in enumerate([18, 34, 12, 10, 24, 22, 24, 22, 18, 18,
                                 20, 20, 14, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for col in (3, 10, 11, 12, 13):
        for row in range(2, ws.max_row + 1):
            ws.cell(row, col).number_format = "#,##0"
    ws.conditional_formatting.add(
        f"N2:N{max(2, ws.max_row)}",
        FormulaRule(formula=['N2="UBICADO"'],
                    fill=PatternFill("solid", fgColor="DCFCE7")))
    ws.conditional_formatting.add(
        f"N2:N{max(2, ws.max_row)}",
        FormulaRule(formula=['N2<>"UBICADO"'],
                    fill=PatternFill("solid", fgColor="FEF3C7")))
    return max(2, ws.max_row)


def _crear_dashboard(wb: Workbook, sku_fin: int, mapa: dict) -> None:
    ws = wb.create_sheet("Dashboard", 1)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    ws["A1"] = "Avance de distribución de localidades"
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL)
    ws["A1"].font = Font(color=BLANCO, bold=True, size=18)
    ws["A1"].alignment = Alignment(vertical="center")
    resumen_ini, resumen_fin = mapa["resumen_inicio"], mapa["resumen_fin"]
    indicadores = [
        ("A4", "SKU por designar", f"=COUNTA('SKU_por_designar'!$A$2:$A${sku_fin})"),
        ("C4", "Tipos de localidad", f"=COUNTA('Mapa_preliminar'!$A${resumen_ini}:$A${resumen_fin})"),
        ("E4", "Localidades requeridas", f"=SUM('Mapa_preliminar'!$B${resumen_ini}:$B${resumen_fin})"),
        ("G4", "Localidades colocadas", f"=SUM('Mapa_preliminar'!$C${resumen_ini}:$C${resumen_fin})"),
        ("A8", "Localidades pendientes", f'=SUMIF(\'Mapa_preliminar\'!$D${resumen_ini}:$D${resumen_fin},">0")'),
        ("C8", "Cobertura del mapa", "=IF(E5=0,0,G5/E5)"),
        ("E8", "Tipos con exceso", f'=COUNTIF(\'Mapa_preliminar\'!$D${resumen_ini}:$D${resumen_fin},"<0")'),
        ("G8", "Asignación de SKU", '="Automática al reimportar"'),
    ]
    for celda, titulo, formula in indicadores:
        col = ws[celda].column
        row = ws[celda].row
        ws.cell(row, col, titulo)
        ws.cell(row, col).font = Font(bold=True, color=AZUL)
        ws.cell(row + 1, col, formula)
        ws.cell(row + 1, col).font = Font(bold=True, size=18, color=AZUL_2)
        ws.cell(row + 1, col).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.cell(row + 1, col).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=col, end_row=row,
                       end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col,
                       end_row=row + 1, end_column=col + 1)
    ws["C9"].number_format = "0%"
    ws.merge_cells("A12:H13")
    ws["A12"] = ("Distribuye tipos en Mapa_preliminar. Después reimporta el libro: "
                   "la aplicación asigna SKU, valida restricciones y regenera "
                   "Mapa_restringido con las dimensiones físicas reales.")
    ws["A12"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A12"].fill = PatternFill("solid", fgColor=AMARILLO)
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 16


def _crear_listas_y_validaciones(wb: Workbook, df, tipos: list[dict], zonas: list[dict]) -> None:
    ws = wb.create_sheet("Listas")
    columnas = {
        "A": ["ABC", "", "A", "B", "C", "D", "E"],
        "B": ["Departamento", ""] + sorted(set(
            df.get("departamento", []).dropna().astype(str))) if hasattr(df.get("departamento", []), "dropna") else ["Departamento", ""],
        "C": ["Clase", ""] + sorted(set(
            df.get("clase_comercial", []).dropna().astype(str))) if hasattr(df.get("clase_comercial", []), "dropna") else ["Clase", ""],
        "D": ["Familia", ""] + sorted(set(
            df.get("familia", []).dropna().astype(str))) if hasattr(df.get("familia", []), "dropna") else ["Familia", ""],
        "E": ["Zona física", ""] + sorted(set(
            df.get("zona_fisica", []).dropna().astype(str))) if hasattr(df.get("zona_fisica", []), "dropna") else ["Zona física", ""],
        "F": ["Tipo"] + [str(t.get("codigo")) for t in tipos],
        "G": ["Zona"] + [str(z.get("nombre")) for z in zonas],
        "H": ["Sí/No", True, False],
        "I": ["Orientación", "", "horizontal", "vertical"],
        "J": ["SKU", ""] + sorted(set(
            df.get("sku", []).dropna().astype(str))) if hasattr(df.get("sku", []), "dropna") else ["SKU", ""],
        "K": (["Código para mapa"]
              + [str(t.get("codigo")) for t in tipos]
              + [f"{t.get('codigo')}|V" for t in tipos]),
    }
    for col, valores in columnas.items():
        for row, valor in enumerate(valores, 1):
            ws[f"{col}{row}"] = valor
    ws.sheet_state = "hidden"
    loc = wb["Localidades"]
    max_row = max(500, loc.max_row + 100)
    validaciones = [
        ("C", "F", len(columnas["F"])), ("D", "G", len(columnas["G"])),
        ("J", "I", len(columnas["I"])), ("K", "A", len(columnas["A"])),
        ("L", "B", len(columnas["B"])), ("M", "C", len(columnas["C"])),
        ("N", "D", len(columnas["D"])), ("O", "E", len(columnas["E"])),
        ("P", "H", len(columnas["H"])), ("Q", "H", len(columnas["H"])),
    ]
    for destino, origen, n in validaciones:
        dv = DataValidation(type="list", formula1=f"'Listas'!${origen}$2:${origen}${max(2, n)}",
                            allow_blank=True)
        loc.add_data_validation(dv)
        dv.add(f"{destino}2:{destino}{max_row}")
    loc.conditional_formatting.add(
        f"A2:S{max_row}",
        FormulaRule(formula=["$Q2=FALSE"], fill=PatternFill("solid", fgColor="E5E7EB")))
    relacion = wb["Relacion_SKU_Localidad"]
    relacion_fin = max(500, relacion.max_row + 100)
    dv_sku = DataValidation(type="list",
                            formula1=f"'Listas'!$J$2:$J${max(2, len(columnas['J']))}",
                            allow_blank=True)
    relacion.add_data_validation(dv_sku)
    dv_sku.add(f"E2:E{relacion_fin}")


def _crear_validacion(wb: Workbook, validacion: dict | None,
                      localidades_pendientes: int = 0) -> None:
    ws = wb.create_sheet("Validacion")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Validación del layout exportado"
    ws["A1"].font = Font(size=16, bold=True, color=AZUL)
    estado = ("REQUIERE CORRECCIÓN"
              if not (validacion or {}).get("valido", True)
              else "PREPARACIÓN EN CURSO" if localidades_pendientes
              else "LISTO")
    ws["A3"], ws["B3"] = "Estado", estado
    ws["A4"], ws["B4"] = "Errores", int((validacion or {}).get("errores", 0))
    ws["A5"], ws["B5"] = "Advertencias", int((validacion or {}).get("advertencias", 0))
    ws["A6"], ws["B6"] = "Localidades pendientes", localidades_pendientes
    color_estado = ("FEE2E2" if estado == "REQUIERE CORRECCIÓN"
                    else "FEF3C7" if localidades_pendientes else "DCFCE7")
    ws["B3"].fill = PatternFill("solid", fgColor=color_estado)
    headers = ["nivel", "codigo", "elemento", "mensaje"]
    _encabezado(ws, 7, headers)
    for issue in (validacion or {}).get("issues", []):
        ws.append([issue.get(c, "") for c in headers])
        color = "FEE2E2" if issue.get("nivel") == "ERROR" else "FEF3C7"
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill("solid", fgColor=color)
    _tabla_estilo(ws, 7, ws.max_row, len(headers))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 72
    ws.protection.sheet = True
    ws.protection.enable()


def _poligono_zona(zona: dict) -> list:
    poly = normalizar_poligono(zona.get("poligono"))
    if poly:
        return poly
    if all(zona.get(k) is not None for k in ("x", "y", "w", "d")):
        x, y = float(zona["x"]), float(zona["y"])
        w, d = float(zona["w"]), float(zona["d"])
        return [(x, y), (x + w, y), (x + w, y + d), (x, y + d)]
    return []


def _punto_en_rectangulo(x: float, y: float, item: dict) -> bool:
    if not all(item.get(k) is not None for k in ("x", "y", "w", "d")):
        return False
    return (float(item["x"]) <= x < float(item["x"]) + float(item["w"])
            and float(item["y"]) <= y < float(item["y"]) + float(item["d"]))


def _crear_mapa_preliminar(wb: Workbook, slots: list[dict], zonas: list[dict],
                           tipos: list[dict], ancho_m: float, largo_m: float,
                           escala_m: float, relacion_fin: int,
                           validacion: dict | None = None, perimetro=None,
                           obstaculos: list[dict] | None = None,
                           accesos: list[dict] | None = None) -> dict:
    """Crea el único editor visible: una ancla por localidad colocada."""
    ws = wb.create_sheet("Mapa_preliminar")
    ws.sheet_view.showGridLines = False
    escala = max(0.2, float(escala_m))
    nx = min(800, max(1, int(math.ceil(float(ancho_m) / escala))))
    ny = min(800, max(1, int(math.ceil(float(largo_m) / escala))))
    codigos = [str(t.get("codigo") or "").strip() for t in tipos
               if str(t.get("codigo") or "").strip()]
    resumen_ini = 5
    resumen_fin = max(resumen_ini, resumen_ini + len(codigos) - 1)
    fila_coord = max(12, resumen_fin + 3)
    fila_inicio, fila_fin = fila_coord + 1, fila_coord + ny
    col_fin = nx + 1
    letra_fin = get_column_letter(col_fin)
    rango_mapa = f"$B${fila_inicio}:${letra_fin}${fila_fin}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=min(max(6, col_fin), 12))
    ws["A1"] = "Mapa preliminar · boceto editable"
    ws["A1"].font = Font(bold=True, size=17, color=BLANCO)
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL)
    ws["A2"] = (f"Cada celda representa una ancla aproximada de {escala:g} m. "
                 "Escribe el tipo (por ejemplo T01) o T01|V para girarlo. "
                 "El SKU compatible se asignará al reimportar; consulta el "
                 "tamaño físico confirmado en Mapa_restringido.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=min(max(6, col_fin), 12))
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 32
    headers = ["tipo_codigo", "requeridas", "colocadas", "restantes",
               "siguiente_localidad", "siguiente_SKU_compatible"]
    _encabezado(ws, 4, headers)
    loc_fin = max(500, len(slots) + 101)
    for i, codigo in enumerate(codigos, resumen_ini):
        ws.cell(i, 1, codigo)
        ws.cell(i, 2, f'=COUNTIF(\'Localidades\'!$C$2:$C${loc_fin},A{i})')
        ws.cell(i, 3, f'=COUNTIF({rango_mapa},A{i})+COUNTIF({rango_mapa},A{i}&"|V")')
        ws.cell(i, 4, f"=B{i}-C{i}")
        ws.cell(i, 5, f'=IF(D{i}<=0,"COMPLETO",IFERROR(INDEX(\'Relacion_SKU_Localidad\'!$A$2:$A${relacion_fin},MATCH(A{i},\'Relacion_SKU_Localidad\'!$C$2:$C${relacion_fin},0)+C{i}),"POR GENERAR"))')
        ws.cell(i, 6, f'=IF(D{i}<=0,"COMPLETO",IFERROR(INDEX(\'Relacion_SKU_Localidad\'!$E$2:$E${relacion_fin},MATCH(A{i},\'Relacion_SKU_Localidad\'!$C$2:$C${relacion_fin},0)+C{i}),"POR ASIGNAR"))')
    _tabla_estilo(ws, 4, resumen_fin, len(headers))
    ws.conditional_formatting.add(
        f"D{resumen_ini}:D{resumen_fin}",
        FormulaRule(formula=[f"D{resumen_ini}<0"],
                    fill=PatternFill("solid", fgColor="FCA5A5")))
    ws.conditional_formatting.add(
        f"D{resumen_ini}:D{resumen_fin}",
        FormulaRule(formula=[f"D{resumen_ini}=0"],
                    fill=PatternFill("solid", fgColor="A7F3D0")))
    for col, width in enumerate([18, 13, 13, 13, 22, 28], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.cell(fila_coord, 1, "Y \\ X")
    ws.cell(fila_coord, 1).font = Font(bold=True, color=AZUL)
    for ix in range(nx):
        col = ix + 2
        ws.cell(fila_coord, col, round(ix * escala, 2))
        ws.cell(fila_coord, col).font = Font(size=7, color="475569")
        ws.column_dimensions[get_column_letter(col)].width = 2.14
    for iy in range(ny):
        row = fila_inicio + iy
        y = (ny - 1 - iy) * escala
        ws.cell(row, 1, round(y, 2))
        ws.cell(row, 1).font = Font(size=7, color="475569")
        ws.row_dimensions[row].height = 15
    ws.column_dimensions["A"].width = 18

    zona_colores = ["E0F2FE", "EDE9FE", "DCFCE7", "FEF3C7", "DBEAFE", "FCE7F3"]
    zonas_poly = [(z, _poligono_zona(z)) for z in zonas]
    per = normalizar_poligono(perimetro)
    bloqueos = list(obstaculos or []) + list(accesos or [])
    for ix in range(nx):
        x = (ix + .5) * escala
        for gy in range(ny):
            y = (gy + .5) * escala
            row, col = fila_inicio + ny - 1 - gy, 2 + ix
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor="E5E7EB")
            zona_idx = next((i for i, (_, poly) in enumerate(zonas_poly)
                             if poly and punto_en_poligono(x, y, poly)), None)
            permitido = (zona_idx is not None
                         and (not per or punto_en_poligono(x, y, per))
                         and not any(_punto_en_rectangulo(x, y, item)
                                     for item in bloqueos))
            if permitido:
                cell.fill = PatternFill(
                    "solid", fgColor=zona_colores[zona_idx % len(zona_colores)])
                cell.protection = Protection(locked=False)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=7, bold=True, color="0F172A")

    lista_fin = max(2, len(codigos) * 2 + 1)
    dv = DataValidation(type="list", formula1=f"'Listas'!$K$2:$K${lista_fin}",
                        allow_blank=True)
    dv.error = "Selecciona un tipo válido o su variante |V."
    dv.errorTitle = "Tipo de localidad no reconocido"
    dv.prompt = "Selecciona el tipo; agrega |V para girar 90°."
    dv.promptTitle = "Colocar localidad"
    dv.showErrorMessage = True
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"B{fila_inicio}:{letra_fin}{fila_fin}")
    colores_tipo = ["86EFAC", "93C5FD", "C4B5FD", "FDE68A", "F9A8D4", "67E8F9"]
    for i, _codigo in enumerate(codigos, resumen_ini):
        ws.conditional_formatting.add(
            f"B{fila_inicio}:{letra_fin}{fila_fin}",
            FormulaRule(
                formula=[f'OR(B{fila_inicio}=$A${i},B{fila_inicio}=$A${i}&"|V")'],
                fill=PatternFill("solid", fgColor=colores_tipo[(i-resumen_ini) % len(colores_tipo)])))

    for s in slots:
        if not all(s.get(k) is not None for k in ("x", "y", "w", "d")):
            continue
        ix, gy = int(float(s["x"]) // escala), int(float(s["y"]) // escala)
        if 0 <= ix < nx and 0 <= gy < ny:
            cell = ws.cell(fila_inicio + ny - 1 - gy, 2 + ix)
            codigo = str(s.get("tipo_codigo") or "").strip()
            if codigo:
                cell.value = codigo + ("|V" if str(s.get("orientacion") or "").lower() == "vertical" else "")

    ws.freeze_panes = f"B{fila_inicio}"
    ws.auto_filter.ref = f"A4:F{resumen_fin}"
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.enable()
    return {"resumen_inicio": resumen_ini, "resumen_fin": resumen_fin,
            "fila_coord": fila_coord, "fila_inicio": fila_inicio,
            "fila_fin": fila_fin, "col_fin": col_fin,
            "escala": escala}


def _crear_mapa_restringido(wb: Workbook, slots: list[dict], zonas: list[dict],
                            tipos: list[dict], ancho_m: float, largo_m: float,
                            escala_m: float, validacion: dict | None = None,
                            perimetro=None, obstaculos: list[dict] | None = None,
                            accesos: list[dict] | None = None) -> None:
    """Dibuja el último resultado validado con huellas físicas reales."""
    ws = wb.create_sheet("Mapa_restringido")
    ws.sheet_view.showGridLines = False
    escala = max(0.2, float(escala_m))
    nx = min(800, max(1, int(math.ceil(float(ancho_m) / escala))))
    ny = min(800, max(1, int(math.ceil(float(largo_m) / escala))))
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=min(max(6, nx + 1), 12))
    ws["A1"] = "Mapa restringido · resultado físico validado"
    ws["A1"].font = Font(bold=True, size=17, color=BLANCO)
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL)
    completas = [s for s in slots
                 if all(s.get(k) is not None for k in ("x", "y", "w", "d"))]
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=min(max(6, nx + 1), 12))
    ws["A2"] = (
        f"Escala: 1 celda ≈ {escala:g} m · {len(completas):,} localidades. "
        + ("Las huellas muestran ancho y fondo reales."
           if completas else
           "Aún no hay localidades validadas: edita Mapa_preliminar, reimporta y descarga de nuevo."))
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30
    fila_coord, fila_inicio = 3, 4
    for ix in range(nx):
        col = ix + 2
        ws.cell(fila_coord, col, round(ix * escala, 2))
        ws.cell(fila_coord, col).font = Font(size=7, color="475569")
        ws.column_dimensions[get_column_letter(col)].width = 2.14
    for iy in range(ny):
        row = fila_inicio + iy
        ws.cell(row, 1, round((ny - 1 - iy) * escala, 2))
        ws.cell(row, 1).font = Font(size=7, color="475569")
        ws.row_dimensions[row].height = 15
    ws["A3"] = "Y \\ X"
    ws["A3"].font = Font(bold=True, color=AZUL)
    ws.column_dimensions["A"].width = 10

    zona_colores = ["E0F2FE", "EDE9FE", "DCFCE7", "FEF3C7", "DBEAFE", "FCE7F3"]
    zonas_poly = [(z, _poligono_zona(z)) for z in zonas]
    per = normalizar_poligono(perimetro)
    for ix in range(nx):
        x = (ix + .5) * escala
        for gy in range(ny):
            y = (gy + .5) * escala
            cell = ws.cell(fila_inicio + ny - 1 - gy, 2 + ix)
            cell.fill = PatternFill("solid", fgColor="E5E7EB")
            zona_idx = next((i for i, (_, poly) in enumerate(zonas_poly)
                             if poly and punto_en_poligono(x, y, poly)), None)
            if (zona_idx is not None
                    and (not per or punto_en_poligono(x, y, per))):
                cell.fill = PatternFill(
                    "solid", fgColor=zona_colores[zona_idx % len(zona_colores)])
    for elementos, color in ((obstaculos or [], "EF4444"),
                              (accesos or [], "F97316")):
        fill = PatternFill("solid", fgColor=color)
        for item in elementos:
            if not all(item.get(k) is not None for k in ("x", "y", "w", "d")):
                continue
            x0, x1 = float(item["x"]), float(item["x"]) + float(item["w"])
            y0, y1 = float(item["y"]), float(item["y"]) + float(item["d"])
            for ix in range(max(0, int(x0 // escala)),
                            min(nx, int(math.ceil(x1 / escala)))):
                for gy in range(max(0, int(y0 // escala)),
                                min(ny, int(math.ceil(y1 / escala)))):
                    ws.cell(fila_inicio + ny - 1 - gy, 2 + ix).fill = fill

    ids_error = {str(i.get("elemento")) for i in (validacion or {}).get("issues", [])
                 if i.get("nivel") == "ERROR"}
    codigos = [str(t.get("codigo") or "") for t in tipos]
    colores_tipo = ["22C55E", "3B82F6", "8B5CF6", "EAB308", "EC4899", "06B6D4"]
    color_por_tipo = {codigo: colores_tipo[i % len(colores_tipo)]
                      for i, codigo in enumerate(codigos)}
    for s in completas:
        x0, x1 = float(s["x"]), float(s["x"]) + float(s["w"])
        y0, y1 = float(s["y"]), float(s["y"]) + float(s["d"])
        cells = []
        fill = PatternFill(
            "solid", fgColor=("FCA5A5" if str(s.get("id")) in ids_error
                               else color_por_tipo.get(str(s.get("tipo_codigo")), "10B981")))
        for ix in range(max(0, int(x0 // escala)),
                        min(nx, int(math.ceil(x1 / escala)))):
            for gy in range(max(0, int(y0 // escala)),
                            min(ny, int(math.ceil(y1 / escala)))):
                cell = ws.cell(fila_inicio + ny - 1 - gy, 2 + ix)
                cell.fill = fill
                cells.append(cell)
        if cells:
            etiqueta = str(s.get("codigo_wms") or s.get("sku_asignado")
                           or s.get("id") or "")
            cells[0].value = etiqueta[:18]
            cells[0].font = Font(size=7, bold=True, color=BLANCO)

    leyenda = fila_inicio + ny + 2
    ws.cell(leyenda, 1, "Leyenda").font = Font(bold=True, color=AZUL)
    for i, z in enumerate(zonas):
        ws.cell(leyenda + i, 2).fill = PatternFill(
            "solid", fgColor=zona_colores[i % len(zona_colores)])
        ws.cell(leyenda + i, 3, str(z.get("nombre") or f"Zona {i + 1}"))
    base = leyenda + len(zonas)
    for i, codigo in enumerate(codigos):
        ws.cell(base + i, 2).fill = PatternFill(
            "solid", fgColor=color_por_tipo[codigo])
        ws.cell(base + i, 3, codigo)
    ws.freeze_panes = "B4"
    ws.protection.sheet = True
    ws.protection.enable()


def exportar_excel(slots: list[dict], zonas: list[dict], tipos: list[dict], df,
                   ancho_m: float, largo_m: float, escala_m: float = 0.5,
                   validacion: dict | None = None, perimetro=None,
                   obstaculos: list[dict] | None = None,
                   accesos: list[dict] | None = None,
                   requerimientos: list[dict] | None = None,
                   relaciones: list[dict] | None = None) -> bytes:
    """Crea un XLSX editable con vista a escala y tablas reimportables."""
    wb = Workbook()
    _crear_instrucciones(wb, escala_m)
    _crear_tipos(wb, tipos)
    _crear_zonas(wb, zonas)
    _crear_localidades(wb, slots)
    relacion_fin = _crear_relacion(wb, slots, relaciones)
    sku_fin = _crear_sku_designar(wb, requerimientos or [], relacion_fin)
    pendientes_geo = sum(not all(s.get(k) is not None
                                 for k in ("x", "y", "w", "d"))
                         for s in slots)
    _crear_validacion(wb, validacion, pendientes_geo)
    mapa = _crear_mapa_preliminar(
        wb, slots, zonas, tipos, ancho_m, largo_m, escala_m, relacion_fin,
        validacion, perimetro, obstaculos, accesos)
    _crear_mapa_restringido(
        wb, slots, zonas, tipos, ancho_m, largo_m, escala_m, validacion,
        perimetro, obstaculos, accesos)
    _crear_listas_y_validaciones(wb, df, tipos, zonas)
    _crear_dashboard(wb, sku_fin, mapa)
    for nombre in ("Instrucciones", "Zonas", "Localidades",
                   "Relacion_SKU_Localidad", "SKU_por_designar",
                   "Validacion", "Listas"):
        wb[nombre].sheet_state = "hidden"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.active = wb.sheetnames.index("Dashboard")
    salida = io.BytesIO()
    wb.save(salida)
    return salida.getvalue()


def _headers(ws) -> dict[str, int]:
    return {str(c.value).strip(): i for i, c in enumerate(ws[1]) if c.value}


def _importar_mapa_preliminar(wb, pool: list[dict], zonas: list[dict]
                              ) -> tuple[list[dict], list[str], list[str]]:
    """Convierte las anclas tipo / tipo|V del mapa en localidades métricas."""
    ws = wb["Mapa_preliminar"]
    fila_coord = next((r for r in range(1, min(ws.max_row, 1000) + 1)
                       if str(ws.cell(r, 1).value or "").strip() == "Y \\ X"), None)
    if not fila_coord:
        return [], ["Mapa_preliminar: no se encontró el origen de la rejilla."], []
    xs = []
    col = 2
    while col <= ws.max_column and isinstance(ws.cell(fila_coord, col).value,
                                               (int, float)):
        xs.append((col, float(ws.cell(fila_coord, col).value)))
        col += 1
    ys = []
    row = fila_coord + 1
    while row <= ws.max_row and isinstance(ws.cell(row, 1).value, (int, float)):
        ys.append((row, float(ws.cell(row, 1).value)))
        row += 1
    if not xs or not ys:
        return [], ["Mapa_preliminar: la rejilla no tiene coordenadas válidas."], []

    tipos = {}
    if "Tipos_ubicacion" in wb.sheetnames:
        wt = wb["Tipos_ubicacion"]
        ht = _headers(wt)
        for fila in wt.iter_rows(min_row=2, values_only=True):
            get = lambda nombre: fila[ht[nombre]] if nombre in ht and ht[nombre] < len(fila) else None
            codigo = str(get("tipo_codigo") or "").strip()
            if not codigo:
                continue
            try:
                tipos[codigo] = {
                    "w": float(get("ancho_m")), "d": float(get("fondo_m")),
                    "h": (float(get("alto_util_m"))
                          if get("alto_util_m") not in (None, "") else None),
                    "tipo_estructura": str(get("estructura") or "PISO").strip(),
                }
            except (TypeError, ValueError):
                continue
    por_tipo: dict[str, list[dict]] = {}
    for item in sorted(pool, key=lambda s: (str(s.get("tipo_codigo") or ""),
                                            str(s.get("id") or ""))):
        por_tipo.setdefault(str(item.get("tipo_codigo") or ""), []).append(item)
    usados_tipo: dict[str, int] = {}
    usados_ids: set[str] = set()
    zonas_poly = [(z, _poligono_zona(z)) for z in zonas]
    slots, errores = [], []
    for row, y in ys:
        for col, x in xs:
            texto = str(ws.cell(row, col).value or "").strip()
            if not texto:
                continue
            vertical = texto.upper().endswith("|V")
            codigo = texto[:-2].strip() if vertical else texto
            tipo = tipos.get(codigo)
            if not tipo:
                errores.append(
                    f"Mapa_preliminar {ws.cell(row, col).coordinate}: "
                    f"tipo desconocido '{texto}'.")
                continue
            indice = usados_tipo.get(codigo, 0)
            candidatos = por_tipo.get(codigo, [])
            if indice >= len(candidatos):
                errores.append(
                    f"Mapa_preliminar: colocaste más localidades '{codigo}' "
                    "de las requeridas.")
                continue
            usados_tipo[codigo] = indice + 1
            s = dict(candidatos[indice])
            usados_ids.add(str(s.get("id")))
            w, d = tipo["w"], tipo["d"]
            if vertical:
                w, d = d, w
            s.update({"x": x, "y": y, "w": w, "d": d,
                      "tipo_codigo": codigo,
                      "orientacion": "vertical" if vertical else "horizontal",
                      "altura_util_nivel_m": tipo.get("h"),
                      "tipo_estructura": tipo.get("tipo_estructura")})
            zona = next((z for z, poly in zonas_poly
                         if poly and rectangulo_en_poligono(s, poly)), None)
            if zona:
                permitidos = {str(v).strip() for v in zona.get("tipos", [])
                              if str(v).strip()}
                if permitidos and codigo not in permitidos:
                    errores.append(
                        f"{s['id']}: el tipo '{codigo}' no está permitido en "
                        f"'{zona.get('nombre')}'.")
                s.update({
                    "zona_layout": zona.get("nombre"),
                    "prioridad": zona.get("prioridad"),
                    "clase_abc_reservada": (zona.get("abc")
                                             or s.get("clase_abc_reservada")),
                    "departamento_reservado": (zona.get("departamentos")
                                                 or s.get("departamento_reservado")),
                    "clase_comercial_reservada": (zona.get("clases")
                                                    or s.get("clase_comercial_reservada")),
                    "familia_reservada": (zona.get("familias")
                                            or s.get("familia_reservada")),
                    "zona_fisica_reservada": ([zona.get("zona_fisica")]
                                               if zona.get("zona_fisica") else None),
                })
            else:
                s["zona_layout"] = None
            slots.append(s)
    pendientes = [str(s.get("id")) for s in pool
                  if str(s.get("id")) not in usados_ids]
    return slots, errores, pendientes


def importar_excel(datos: bytes) -> dict:
    """Recupera localidades y zonas editadas en el libro de intercambio."""
    try:
        wb = load_workbook(io.BytesIO(datos), data_only=False)
    except Exception as exc:
        return {"slots": [], "zonas": [], "errores": [f"No se pudo abrir el XLSX: {exc}"]}
    errores, slots, zonas, relaciones, pendientes, pool = [], [], [], [], [], []
    if "Localidades" not in wb.sheetnames:
        return {"slots": [], "zonas": [],
                "errores": ["Falta la hoja obligatoria 'Localidades'."]}
    ws = wb["Localidades"]
    h = _headers(ws)
    faltantes = [c for c in ("id_localidad", "x_m", "y_m", "ancho_m", "fondo_m")
                 if c not in h]
    if faltantes:
        return {"slots": [], "zonas": [],
                "errores": ["Faltan columnas: " + ", ".join(faltantes)]}
    ids = set()
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(v not in (None, "") for v in row):
            continue
        get = lambda nombre, default=None: row[h[nombre]] if nombre in h and h[nombre] < len(row) else default
        uid = str(get("id_localidad") or "").strip()
        if not uid:
            errores.append(f"Fila {n}: id_localidad vacío.")
            continue
        if uid in ids:
            errores.append(f"Fila {n}: id_localidad duplicado '{uid}'.")
            continue
        ids.add(uid)
        reserva_zf = [p.strip() for p in str(
            get("zona_fisica_permitida") or "").replace(";", ",").split(",")
            if p.strip()]
        base = {
            "id": uid, "codigo_wms": str(get("codigo_wms") or "").strip() or None,
            "tipo_codigo": str(get("tipo_codigo") or "").strip() or None,
            "zona_layout": str(get("zona_layout") or "").strip() or None,
            "altura_util_nivel_m": (float(get("alto_util_m"))
                                      if get("alto_util_m") not in (None, "") else None),
            "orientacion": str(get("orientacion") or "").strip() or None,
            "clase_abc_reservada": str(get("abc_permitido") or "").strip() or None,
            "departamento_reservado": str(get("departamento_permitido") or "").strip() or None,
            "clase_comercial_reservada": str(get("clase_permitida") or "").strip() or None,
            "familia_reservada": str(get("familia_permitida") or "").strip() or None,
            "zona_fisica_reservada": reserva_zf or None,
            "multisku": _booleano(get("multisku")),
            "activa": _booleano(get("activa"), True),
            "notas": str(get("notas") or "").strip() or None,
            "niveles": None, "prioridad": None,
        }
        for destino, origen in (("w", "ancho_m"), ("d", "fondo_m")):
            try:
                base[destino] = float(get(origen))
            except (TypeError, ValueError):
                base[destino] = None
        pool.append(base)
        valores_geo = [get("x_m"), get("y_m"), get("ancho_m"), get("fondo_m")]
        if any(v in (None, "") for v in valores_geo):
            pendientes.append(uid)
            continue
        try:
            x, y = float(get("x_m")), float(get("y_m"))
            w, d = float(get("ancho_m")), float(get("fondo_m"))
            if min(x, y) < 0 or min(w, d) <= 0:
                raise ValueError
        except (TypeError, ValueError):
            errores.append(f"Fila {n}: coordenadas o dimensiones inválidas.")
            continue
        slots.append({**base, "x": x, "y": y, "w": w, "d": d})
    if "Zonas" in wb.sheetnames:
        wz, hz = wb["Zonas"], _headers(wb["Zonas"])
        for n, row in enumerate(wz.iter_rows(min_row=2, values_only=True), 2):
            get = lambda nombre, default=None: row[hz[nombre]] if nombre in hz and hz[nombre] < len(row) else default
            nombre = str(get("zona") or "").strip()
            if not nombre:
                continue
            z = {"nombre": nombre, "prioridad": int(get("prioridad") or len(zonas) + 1)}
            vertices = str(get("vertices_json") or "").strip()
            if vertices:
                try:
                    z["poligono"] = json.loads(vertices)
                except json.JSONDecodeError:
                    errores.append(f"Zonas fila {n}: vertices_json inválido.")
                    continue
            else:
                try:
                    z.update({"x": float(get("x_m")), "y": float(get("y_m")),
                              "w": float(get("ancho_m")), "d": float(get("fondo_m"))})
                except (TypeError, ValueError):
                    errores.append(f"Zonas fila {n}: geometría inválida.")
                    continue
            for campo, origen in (("zona_fisica", "zona_fisica"),
                                  ("tipo_estructura", "estructura"),
                                  ("modo_pasillo", "pasillos"),
                                  ("pasillo_m", "ancho_pasillo_m"),
                                  ("orientacion", "orientacion"),
                                  ("margen_m", "margen_m")):
                valor = get(origen)
                if valor not in (None, ""):
                    z[campo] = valor
            for destino, origen in (("tipos", "tipos_permitidos"),
                                    ("departamentos", "departamentos"),
                                    ("clases", "clases"), ("familias", "familias"),
                                    ("abc", "abc")):
                z[destino] = [p.strip() for p in str(get(origen) or "").replace(";", ",").split(",") if p.strip()]
            zonas.append(z)
    modo_mapa = "Mapa_preliminar" in wb.sheetnames
    if modo_mapa:
        slots_mapa, errores_mapa, pendientes_mapa = _importar_mapa_preliminar(
            wb, pool, zonas)
        slots = slots_mapa
        pendientes = pendientes_mapa
        errores.extend(errores_mapa)
    if "Relacion_SKU_Localidad" in wb.sheetnames and not modo_mapa:
        wr = wb["Relacion_SKU_Localidad"]
        hr = _headers(wr)
        for row in wr.iter_rows(min_row=2, values_only=False):
            def valor(nombre, default=None):
                return (row[hr[nombre]].value
                        if nombre in hr and hr[nombre] < len(row) else default)
            uid = str(valor("id_localidad") or "").strip()
            sku = str(valor("sku_asignado") or "").strip()
            if uid and sku:
                relaciones.append({"id_localidad": uid, "sku": sku})
    return {"slots": slots, "zonas": zonas, "errores": errores,
            "relaciones": relaciones,
            "localidades_pendientes": pendientes}


__all__ = ["exportar_excel", "importar_excel", "LOCALIDAD_COLS"]
