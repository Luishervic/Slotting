"""Contrato de ida y vuelta del editor masivo de layout."""
from __future__ import annotations

import io
import unittest
import xml.etree.ElementTree as ET
import zipfile

import pandas as pd
from openpyxl import load_workbook

from slotting import layout_exchange as LX


class TestLayoutExchange(unittest.TestCase):
    def setUp(self):
        self.df = pd.DataFrame({
            "sku": ["A", "B"], "departamento": ["MOTOS", "PISO"],
            "clase_comercial": ["C1", "C2"], "familia": ["F1", "F2"],
            "zona_fisica": ["PISO", "RACK"], "clase_abc": ["A", "C"],
        })
        self.tipos = [{"codigo": "PIS-T01", "talla": "Compacta",
                       "zona_fisica": "PISO", "tipo_estructura": "PISO",
                       "w": 1.2, "d": .8, "h": 1.5, "n_skus": 2}]
        self.zonas = [{"nombre": "Zona norte", "x": 0, "y": 0, "w": 10,
                       "d": 8, "prioridad": 1, "orientacion": "horizontal",
                       "departamentos": ["MOTOS"], "tipo_estructura": "PISO",
                       "modo_pasillo": "con", "pasillo_m": 3.5}]
        self.slots = [{"id": "U-001", "codigo_wms": "WMS-01",
                       "tipo_codigo": "PIS-T01", "zona_layout": "Zona norte",
                       "x": 1.0, "y": 2.0, "w": 1.2, "d": .8,
                       "altura_util_nivel_m": 1.5,
                       "departamento_reservado": "MOTOS",
                       "clase_abc_reservada": "A", "multisku": False}]
        self.requerimientos = [
            {"sku": "A", "descripcion": "Moto", "unidades": 10,
             "abc": "A", "departamento": "MOTOS", "clase": "C1",
             "familia": "F1", "zona_fisica": "PISO",
             "tipo_codigo": "PIS-T01", "capacidad_localidad": 5,
             "localidades_necesarias": 2},
            {"sku": "B", "descripcion": "Piso", "unidades": 4,
             "abc": "C", "departamento": "PISO", "clase": "C2",
             "familia": "F2", "zona_fisica": "RACK",
             "tipo_codigo": "PIS-T01", "capacidad_localidad": 4,
             "localidades_necesarias": 1},
        ]

    def test_libro_separa_boceto_editable_y_resultado_restringido(self):
        datos = LX.exportar_excel(
            self.slots, self.zonas, self.tipos, self.df, 10, 8, .5,
            requerimientos=self.requerimientos)
        wb = load_workbook(io.BytesIO(datos))
        visibles = [ws.title for ws in wb.worksheets
                    if ws.sheet_state == "visible"]
        self.assertEqual(visibles, [
            "Dashboard", "Tipos_ubicacion", "Mapa_preliminar",
            "Mapa_restringido"])
        self.assertEqual(wb["Zonas"].sheet_state, "hidden")
        self.assertEqual(wb["Localidades"].sheet_state, "hidden")
        mapa = wb["Mapa_preliminar"]
        self.assertTrue(mapa.protection.sheet)
        self.assertIn("asignará", mapa["A2"].value)
        restringido = wb["Mapa_restringido"]
        self.assertTrue(restringido.protection.sheet)
        self.assertIn("huellas muestran ancho y fondo reales",
                      restringido["A2"].value)
        celdas_tipo = sum(
            c.fill.fgColor.rgb in {"0022C55E", "22C55E"}
            for row in restringido.iter_rows() for c in row)
        self.assertGreaterEqual(celdas_tipo, 6)
        self.assertEqual(list(wb["Localidades"].tables), ["tblLocalidades"])
        self.assertEqual(list(wb["Relacion_SKU_Localidad"].tables),
                         ["tblRelacionSkuLocalidad"])
        self.assertEqual(list(wb["SKU_por_designar"].tables),
                         ["tblSkuPorDesignar"])
        self.assertTrue(wb["Dashboard"]["G5"].value.startswith("=SUM"))
        self.assertIn("COUNTIF", mapa["C5"].value)
        self.assertEqual(mapa["D5"].value, "=B5-C5")
        fila_coord = next(r for r in range(1, mapa.max_row + 1)
                          if mapa.cell(r, 1).value == "Y \\ X")
        self.assertFalse(mapa.cell(fila_coord + 1, 2).protection.locked)
        self.assertTrue(mapa.data_validations.count)
        self.assertIn("COUNTIFS", wb["SKU_por_designar"]["L2"].value)
        self.assertEqual(wb["Localidades"]["S2"].value,
                         '=IF(OR(A2="",C2="",D2="",E2="",F2="",G2="",H2=""),"PENDIENTE","PREPARADA")')

    def test_tablas_no_superponen_autofiltro_de_hoja(self):
        datos = LX.exportar_excel(
            self.slots, self.zonas, self.tipos, self.df, 10, 8, .5,
            requerimientos=self.requerimientos)
        ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        with zipfile.ZipFile(io.BytesIO(datos)) as paquete:
            tablas = [nombre for nombre in paquete.namelist()
                      if nombre.startswith("xl/tables/table")]
            self.assertEqual(len(tablas), 5)
            for nombre in paquete.namelist():
                if not (nombre.startswith("xl/worksheets/sheet")
                        and nombre.endswith(".xml")):
                    continue
                raiz = ET.fromstring(paquete.read(nombre))
                tiene_tabla = raiz.find(f"{ns}tableParts") is not None
                tiene_filtro_hoja = raiz.find(f"{ns}autoFilter") is not None
                self.assertFalse(
                    tiene_tabla and tiene_filtro_hoja,
                    f"{nombre} contiene tabla y AutoFilter de hoja superpuestos")

    def test_ida_y_vuelta_preserva_geometria_y_restricciones(self):
        datos = LX.exportar_excel(self.slots, self.zonas, self.tipos, self.df,
                                  10, 8, .5)
        out = LX.importar_excel(datos)
        self.assertEqual(out["errores"], [])
        self.assertEqual(out["slots"][0]["codigo_wms"], "WMS-01")
        self.assertEqual(out["slots"][0]["departamento_reservado"], ["MOTOS"])
        self.assertEqual(out["slots"][0]["clase_abc_reservada"], "A")
        self.assertEqual(out["zonas"][0]["departamentos"], ["MOTOS"])
        self.assertEqual(out["zonas"][0]["tipo_estructura"], "PISO")
        self.assertEqual(out["zonas"][0]["modo_pasillo"], "con")
        self.assertEqual(out["zonas"][0]["pasillo_m"], 3.5)

    def test_detecta_ids_duplicados(self):
        datos = LX.exportar_excel(self.slots, self.zonas, self.tipos, self.df,
                                  10, 8, .5)
        wb = load_workbook(io.BytesIO(datos))
        ws = wb["Localidades"]
        for c in range(1, ws.max_column + 1):
            ws.cell(3, c, ws.cell(2, c).value)
        salida = io.BytesIO()
        wb.save(salida)
        out = LX.importar_excel(salida.getvalue())
        self.assertTrue(any("duplicado" in e for e in out["errores"]))

    def test_importa_relaciones_y_tolera_geometria_pendiente(self):
        pendientes = [{**self.slots[0], "x": None, "y": None}]
        datos = LX.exportar_excel(
            pendientes, self.zonas, self.tipos, self.df, 10, 8, .5,
            requerimientos=self.requerimientos,
            relaciones=[{"id_localidad": "U-001", "sku": "A"}])
        out = LX.importar_excel(datos)
        self.assertEqual(out["errores"], [])
        self.assertEqual(out["slots"], [])
        self.assertEqual(out["localidades_pendientes"], ["U-001"])
        self.assertEqual(out["relaciones"], [])

    def test_mapa_define_geometria_y_rotacion(self):
        pendientes = [{**self.slots[0], "x": None, "y": None}]
        datos = LX.exportar_excel(
            pendientes, self.zonas, self.tipos, self.df, 10, 8, .5,
            requerimientos=self.requerimientos)
        wb = load_workbook(io.BytesIO(datos))
        ws = wb["Mapa_preliminar"]
        fila_coord = next(r for r in range(1, ws.max_row + 1)
                          if ws.cell(r, 1).value == "Y \\ X")
        col = next(c for c in range(2, ws.max_column + 1)
                   if ws.cell(fila_coord, c).value == 1.0)
        row = next(r for r in range(fila_coord + 1, ws.max_row + 1)
                   if ws.cell(r, 1).value == 2.0)
        ws.cell(row, col, "PIS-T01|V")
        salida = io.BytesIO()
        wb.save(salida)
        out = LX.importar_excel(salida.getvalue())
        self.assertEqual(out["errores"], [])
        self.assertEqual(len(out["slots"]), 1)
        slot = out["slots"][0]
        self.assertEqual((slot["x"], slot["y"]), (1.0, 2.0))
        self.assertEqual((slot["w"], slot["d"]), (.8, 1.2))
        self.assertEqual(slot["orientacion"], "vertical")
        self.assertEqual(slot["zona_layout"], "Zona norte")


if __name__ == "__main__":
    unittest.main()
